import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk
import openpyxl
from openpyxl.utils.cell import range_boundaries
import re
import os
import threading
import queue
import datetime
import logging
import json
import csv
import multiprocessing
import platform
from functools import partial

# --- Logic-Only Worker Function ---
def process_single_file_task(file_path, sheet_name, cell_range, mode, ignore_words, ignore_chars, skip_hidden, is_report=False):
    try:
        is_csv = file_path.lower().endswith('.csv')
        results = {"file": os.path.basename(file_path), "success": True, "sheets": []}
        file_words, file_chars = 0, 0
        if is_csv:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f)
                sheet_words, sheet_chars = 0, 0
                for row in reader:
                    for cell_val in row:
                        if not cell_val: continue
                        s_val = str(cell_val).strip()
                        if s_val.replace('.', '', 1).isdigit(): continue
                        if ignore_chars:
                            for c in ignore_chars: s_val = s_val.replace(c, '')
                        if mode in ["Word", "Both"]:
                            words = [w for w in s_val.split() if w.lower() not in ignore_words]
                            sheet_words += len(words)
                        if mode in ["Char", "Both"]:
                            sheet_chars += len(re.sub(r'\s+', '', s_val))
                results["sheets"].append({"name": "CSV Data", "words": sheet_words, "chars": sheet_chars})
                file_words, file_chars = sheet_words, sheet_chars
        else:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            target_sheets = wb.sheetnames if is_report else ([sheet_name] if sheet_name in wb.sheetnames else [])
            for s_name in target_sheets:
                ws = wb[s_name]
                sheet_words, sheet_chars = 0, 0
                try: dim = ws.calculate_dimension() if is_report else cell_range
                except: dim = "A1:A1"
                min_col, min_row, max_col, max_row = range_boundaries(dim)
                if max_row is None: max_row = ws.max_row
                hidden_cols = set()
                if skip_hidden:
                    from openpyxl.utils import get_column_letter
                    for c in range(min_col, max_col + 1):
                        if c in ws.column_dimensions and ws.column_dimensions[get_column_letter(c)].hidden:
                            hidden_cols.add(c)
                for r_idx, row in enumerate(ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col), start=min_row):
                    if skip_hidden and r_idx in ws.row_dimensions and ws.row_dimensions[r_idx].hidden: continue
                    for cell in row:
                        if cell.value is None: continue
                        if skip_hidden and cell.column in hidden_cols: continue
                        s_val = str(cell.value).strip()
                        if s_val.replace('.', '', 1).isdigit(): continue
                        if ignore_chars:
                            for c in ignore_chars: s_val = s_val.replace(c, '')
                        if mode in ["Word", "Both"]:
                            words = [w for w in s_val.split() if w.lower() not in ignore_words]
                            sheet_words += len(words)
                        if mode in ["Char", "Both"]:
                            sheet_chars += len(re.sub(r'\s+', '', s_val))
                results["sheets"].append({"name": s_name, "words": sheet_words, "chars": sheet_chars})
                file_words += sheet_words; file_chars += sheet_chars
            wb.close()
        results["total_words"], results["total_chars"] = file_words, file_chars
        return results
    except Exception as e:
        return {"file": os.path.basename(file_path), "success": False, "error": str(e)}

CONFIG_FILE = "settings.json"
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(message)s', handlers=[logging.FileHandler("app_debug.log"), logging.StreamHandler()])
logger = logging.getLogger(__name__)

def validate_range(range_str):
    if not range_str: return False
    return bool(re.match(r'^([A-Z]+[0-9]+:[A-Z]+[0-9]+|[A-Z]+:[A-Z]+|[0-9]+:[0-9]+)$', range_str.upper()))

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    HAS_DND = True
except ImportError:
    HAS_DND = False

class SettingsWindow(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.title("Settings")
        self.geometry("400x480")
        self.after(10, self.lift)
        self.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() // 2) - 200
        y = parent.winfo_y() + (parent.winfo_height() // 2) - 240
        self.geometry(f"+{x}+{y}")
        frame = ctk.CTkFrame(self, fg_color="transparent"); frame.pack(fill="both", expand=True, padx=30, pady=20)
        ctk.CTkLabel(frame, text="Appearance Mode", font=parent.heading_font, text_color=parent.text_primary).pack(anchor="w")
        self.theme_menu = ctk.CTkOptionMenu(frame, values=["Light", "Dark"], font=parent.ui_font, fg_color=parent.btn_bg, button_color=parent.btn_bg, text_color=parent.text_primary, command=parent.change_appearance)
        self.theme_menu.pack(fill="x", pady=(5, 20)); self.theme_menu.set(parent.appearance_mode.capitalize())
        ctk.CTkLabel(frame, text="Ignore Words", font=parent.heading_font).pack(anchor="w")
        self.entry_words = ctk.CTkEntry(frame, font=parent.ui_font, fg_color=parent.btn_bg, border_color=parent.border_color, text_color=parent.text_primary); self.entry_words.pack(fill="x", pady=(5, 20)); self.entry_words.insert(0, parent.ignore_words_val)
        ctk.CTkLabel(frame, text="Ignore Characters", font=parent.heading_font).pack(anchor="w")
        self.entry_chars = ctk.CTkEntry(frame, font=parent.ui_font, fg_color=parent.btn_bg, border_color=parent.border_color, text_color=parent.text_primary); self.entry_chars.pack(fill="x", pady=(5, 20)); self.entry_chars.insert(0, parent.ignore_chars_val)
        self.check_hidden = ctk.CTkCheckBox(frame, text="Skip hidden rows & columns", font=parent.ui_font, text_color=parent.text_primary, fg_color=parent.accent_color); self.check_hidden.pack(anchor="w", pady=(0, 10))
        if parent.skip_hidden_val: self.check_hidden.select()
        self.check_parallel = ctk.CTkCheckBox(frame, text="Fast Parallel Processing", font=parent.ui_font, text_color=parent.text_primary, fg_color=parent.accent_color); self.check_parallel.pack(anchor="w", pady=(0, 20))
        if parent.parallel_val: self.check_parallel.select()
        ctk.CTkButton(frame, text="Apply & Close", font=parent.heading_font, fg_color=parent.accent_color, text_color=parent.btn_text, height=40, command=self.save).pack(fill="x", side="bottom")

    def save(self):
        self.parent.ignore_words_val, self.parent.ignore_chars_val = self.entry_words.get(), self.entry_chars.get()
        self.parent.skip_hidden_val, self.parent.parallel_val = self.check_hidden.get(), self.check_parallel.get()
        self.parent.save_settings(); self.destroy()

class ExcelCounterApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.appearance_mode = "light"
        self.text_primary, self.text_secondary = ("#37352F", "#FFFFFF"), ("#73726E", "#A0A0A0")
        self.border_color, self.log_bg = ("#E9E9E7", "#3F3F3C"), ("#F7F6F3", "#191919")
        self.accent_color, self.btn_text = ("#37352F", "#FFFFFF"), ("#FFFFFF", "#000000")
        self.btn_bg, self.btn_hover = ("#FFFFFF", "#2D2D2A"), ("#000000", "#E0E0E0")
        self.ui_font, self.heading_font, self.radius = ("Inter", 13), ("Inter", 16, "bold"), 6
        self.file_path, self.file_list, self.workbook, self.is_processing = "", [], None, False
        self.msg_queue, self.stop_event, self.settings_window = queue.Queue(), threading.Event(), None
        self.ignore_words_val, self.ignore_chars_val, self.skip_hidden_val, self.parallel_val = "", "", False, True
        self.title("Text Counter"); self.geometry("500x620")
        self.main_frame = ctk.CTkFrame(self, fg_color="transparent"); self.main_frame.pack(fill="both", expand=True)
        if HAS_DND:
            try:
                arch = 'win-x64' if platform.architecture()[0] == '64bit' else 'win-x86'
                dnd_path = os.path.join(os.path.dirname(csv.__file__), '..', 'tkinterdnd2', 'tkdnd', arch) # fallback find
                import tkinterdnd2; dnd_path = os.path.join(os.path.dirname(tkinterdnd2.__file__), 'tkdnd', arch)
                self.tk.call('lappend', 'auto_path', dnd_path); TkinterDnD._require(self)
                self.main_frame.drop_target_register(DND_FILES); self.main_frame.dnd_bind('<<Drop>>', self.handle_drop)
            except: pass
        self.create_widgets(); self.after(100, self.load_settings); self.check_queue()

    def change_appearance(self, m): self.appearance_mode = m.lower(); ctk.set_appearance_mode(m); self.save_settings()
    def open_settings(self):
        if not self.settings_window or not self.settings_window.winfo_exists(): self.settings_window = SettingsWindow(self)
        else: self.settings_window.focus()

    def load_settings(self):
        try:
            if os.path.exists(CONFIG_FILE):
                with open(CONFIG_FILE, "r") as f:
                    s = json.load(f); self.appearance_mode = s.get("appearance_mode", "light"); ctk.set_appearance_mode(self.appearance_mode)
                    lp = s.get("last_file_path", "")
                    if lp and os.path.exists(lp): self.load_file(lp)
                    self.mode_var.set(s.get("last_mode", "Word")); self.entry_range.delete(0, tk.END); self.entry_range.insert(0, s.get("last_range", ""))
                    self.ignore_words_val, self.ignore_chars_val = s.get("ignore_words", ""), s.get("ignore_chars", "")
                    self.skip_hidden_val, self.parallel_val = s.get("skip_hidden", False), s.get("skip_parallel", True)
        except: pass

    def save_settings(self):
        try:
            s = {"appearance_mode": self.appearance_mode, "last_file_path": self.file_path, "last_mode": self.mode_var.get(), "last_range": self.entry_range.get(), "ignore_words": self.ignore_words_val, "ignore_chars": self.ignore_chars_val, "skip_hidden": self.skip_hidden_val, "skip_parallel": self.parallel_val}
            with open(CONFIG_FILE, "w") as f: json.dump(s, f)
        except: pass

    def process_with_granular_progress(self, file_path, sheet_name, cell_range, mode, i_words, i_chars, skip_h):
        """Standard task with row-by-row UI updates for single files."""
        try:
            is_csv = file_path.lower().endswith('.csv'); f_name = os.path.basename(file_path)
            self.msg_queue.put({"type": "progress_raw", "value": 0.05, "file": f"📂 Loading {f_name}..."})
            if is_csv:
                # Use basic sequential processing for simplicity in this version
                res = process_single_file_task(file_path, sheet_name, cell_range, mode, i_words, i_chars, skip_h)
            else:
                res = process_single_file_task(file_path, sheet_name, cell_range, mode, i_words, i_chars, skip_h)
            return res
        except Exception as e: return {"success": False, "error": str(e)}

    def counting_thread_target(self, target_files, sheet_name, cell_range, mode):
        try:
            gw, gc, total = 0, 0, len(target_files)
            i_words = [w.strip().lower() for w in self.ignore_words_val.split(',') if w.strip()]
            i_chars, skip_h = self.ignore_chars_val, self.skip_hidden_val
            if self.parallel_val and total > 1:
                self.msg_queue.put({"type": "progress_raw", "value": 0.1, "file": "🚀 Starting Parallel Engine..."})
                worker = partial(process_single_file_task, sheet_name=sheet_name, cell_range=cell_range, mode=mode, ignore_words=i_words, ignore_chars=i_chars, skip_hidden=skip_h)
                with multiprocessing.Pool(processes=min(multiprocessing.cpu_count(), total)) as pool:
                    for i, res in enumerate(pool.imap_unordered(worker, target_files)):
                        if self.stop_event.is_set(): pool.terminate(); return
                        if res and res.get("success"): gw += res["total_words"]; gc += res["total_chars"]; self.msg_queue.put({"type": "progress", "value": i + 1, "total": total, "file": res['file']})
            else:
                for f_idx, fp in enumerate(target_files):
                    if self.stop_event.is_set(): return
                    self.msg_queue.put({"type": "progress", "value": f_idx, "total": total, "file": os.path.basename(fp)})
                    res = process_single_file_task(fp, sheet_name, cell_range, mode, i_words, i_chars, skip_h)
                    if res and res.get("success"): gw += res["total_words"]; gc += res["total_chars"]
                    self.msg_queue.put({"type": "progress", "value": f_idx + 1, "total": total, "file": os.path.basename(fp)})
            ts = datetime.datetime.now().strftime("%I:%M %p")
            res_str = f"[{ts}] Count Complete\n"
            if mode == "Both": res_str += f"✨ {gw:,} words | {gc:,} chars\n\n"
            else: res_str += f"✨ {gw if mode=='Word' else gc:,} {mode.lower()}s\n\n"
            self.msg_queue.put({"type": "done", "result_str": res_str})
        except Exception as e: self.msg_queue.put({"type": "error", "message": str(e)})

    def report_thread_target(self, target_files, mode):
        try:
            gw, gc, report, all_res = 0, 0, "", []
            i_words = [w.strip().lower() for w in self.ignore_words_val.split(',') if w.strip()]
            i_chars, skip_h = self.ignore_chars_val, self.skip_hidden_val
            if self.parallel_val and len(target_files) > 1:
                self.msg_queue.put({"type": "progress_raw", "value": 0.1, "file": "📊 Running Parallel Report..."})
                worker = partial(process_single_file_task, sheet_name="", cell_range="", mode=mode, ignore_words=i_words, ignore_chars=i_chars, skip_hidden=skip_h, is_report=True)
                with multiprocessing.Pool(processes=min(multiprocessing.cpu_count(), len(target_files))) as pool:
                    for i, res in enumerate(pool.imap(worker, target_files)):
                        if self.stop_event.is_set(): pool.terminate(); return
                        if res: all_res.append(res); self.msg_queue.put({"type": "progress", "value": i + 1, "total": len(target_files), "file": res.get('file', '...')})
            else:
                for f_idx, fp in enumerate(target_files):
                    if self.stop_event.is_set(): return
                    self.msg_queue.put({"type": "progress", "value": f_idx, "total": len(target_files), "file": os.path.basename(fp)})
                    res = process_single_file_task(fp, "", "", mode, i_words, i_chars, skip_h, True)
                    if res: all_res.append(res)
            for res in all_res:
                if not res or not res.get("success"): report += f"❌ {res.get('file','???')}: {res.get('error','Worker died')}\n\n"; continue
                report += f"--- {res['file']} ---\n"
                for s in res["sheets"]:
                    if mode == "Both": report += f" • {s['name']}: {s['words']:,} w / {s['chars']:,} c\n"
                    else: report += f" • {s['name']}: {s['words'] if mode=='Word' else s['chars']:,} {mode.lower()}s\n"
                report += f"➔ File Total: {res['total_words'] if mode=='Word' else res['total_chars']:,} {mode.lower()}s\n\n"
                gw += res["total_words"]; gc += res["total_chars"]
            header = f"[{datetime.datetime.now().strftime('%I:%M %p')}] 📊 FULL REPORT\n"
            if mode == "Both": header += f"✨ Total: {gw:,} w | {gc:,} c\n\n"
            else: header += f"✨ Total: {gw if mode=='Word' else gc:,} {mode.lower()}s\n\n"
            self.msg_queue.put({"type": "done", "result_str": header + report})
        except Exception as e: self.msg_queue.put({"type": "error", "message": str(e)})

    def run_count(self):
        if self.is_processing: return
        target = self.file_list if self.batch_var.get() else ([self.file_path] if self.file_path else [])
        if not target: messagebox.showwarning("Error", "Select file(s)."); return
        rng = self.entry_range.get().strip().upper()
        if not validate_range(rng): messagebox.showwarning("Error", "Invalid Range."); return
        self.is_processing = True; self.stop_event.clear(); self.btn_run.configure(state="disabled"); self.btn_report.configure(state="disabled")
        self.btn_cancel.configure(state="normal"); self.progress.set(0); self.lbl_status.configure(text="Processing...")
        threading.Thread(target=self.counting_thread_target, args=(target, self.combo_sheet.get(), rng, self.mode_var.get()), daemon=True).start()

    def run_report(self):
        if self.is_processing: return
        target = self.file_list if self.batch_var.get() else ([self.file_path] if self.file_path else [])
        if not target: messagebox.showwarning("Error", "Select file(s)."); return
        self.is_processing = True; self.stop_event.clear(); self.btn_run.configure(state="disabled"); self.btn_report.configure(state="disabled")
        self.btn_cancel.configure(state="normal"); self.progress.set(0); self.lbl_status.configure(text="Generating Report...")
        threading.Thread(target=self.report_thread_target, args=(target, self.mode_var.get()), daemon=True).start()

    def cancel_count(self): self.stop_event.set(); self.lbl_status.configure(text="Cancelling...")
    def reset_ui_after_run(self): self.btn_run.configure(state="normal"); self.btn_report.configure(state="normal"); self.btn_cancel.configure(state="disabled"); self.progress.set(0); self.lbl_status.configure(text="✅ Ready"); self.is_processing = False; self.save_settings()
    def clear_log(self): self.out_area.configure(state='normal'); self.out_area.delete("1.0", tk.END); self.out_area.configure(state='disabled')
    def export_log(self):
        c = self.out_area.get("1.0", tk.END).strip()
        if not c: return
        f = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text", "*.txt"), ("CSV", "*.csv")])
        if f:
            try:
                if f.endswith('.csv'): 
                    lines, rows = c.split('\n'), [["Timestamp", "Details", "Result"]]
                    for line in lines:
                        if not line.strip(): continue
                        m = re.search(r'\[(.*?)\]', line)
                        if m: rows.append([m.group(1), line.replace(f"[{m.group(1)}]", "").strip(), ""])
                    with open(f, 'w', newline='', encoding='utf-8') as file: csv.writer(file).writerows(rows)
                else:
                    with open(f, "w", encoding="utf-8") as file: file.write(c)
                messagebox.showinfo("Exported", "Saved successfully!")
            except Exception as e: messagebox.showerror("Error", str(e))

    def browse_file(self):
        ft = [("All", "*.xlsx *.csv"), ("Excel", "*.xlsx"), ("CSV", "*.csv")]
        if self.batch_var.get():
            files = filedialog.askopenfilenames(filetypes=ft)
            if files: self.file_list = list(files); self.lbl_file.configure(text=f"{len(files)} files")
        else:
            f = filedialog.askopenfilename(filetypes=ft); 
            if f: self.load_file(f)

    def load_file(self, f):
        try:
            if f.lower().endswith('.csv'): self.file_path, self.workbook = f, None; self.lbl_file.configure(text=os.path.basename(f)); self.combo_sheet.configure(values=["Default (CSV)"]); self.combo_sheet.set("Default (CSV)")
            else: self.workbook = openpyxl.load_workbook(f, read_only=True, data_only=True); self.file_path = f; self.lbl_file.configure(text=os.path.basename(f)); self.combo_sheet.configure(values=self.workbook.sheetnames); self.combo_sheet.set(self.workbook.sheetnames[0])
        except Exception as e: messagebox.showerror("Error", str(e))

    def handle_drop(self, event):
        if self.is_processing: return
        paths = [p[0] or p[1] for p in re.findall(r'\{(.*?)\}|(\S+)', event.data.strip())]
        supported = [p for p in paths if p.lower().endswith(('.xlsx', '.csv'))]
        if self.batch_var.get(): self.file_list = supported; self.lbl_file.configure(text=f"{len(supported)} files")
        elif supported: self.load_file(supported[0])

    def scan_range(self):
        if not self.file_path: return
        
        # --- CSV Auto-detect ---
        if self.file_path.lower().endswith('.csv'):
            try:
                self.lbl_status.configure(text="🔍 Scanning CSV size...")
                max_cols = 0
                max_rows = 0
                with open(self.file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    reader = csv.reader(f)
                    for i, row in enumerate(reader):
                        if any(cell.strip() for cell in row):
                            max_rows = i + 1
                            max_cols = max(max_cols, len(row))
                
                if max_rows > 0:
                    from openpyxl.utils import get_column_letter
                    dim = f"A1:{get_column_letter(max_cols)}{max_rows}"
                    self.entry_range.delete(0, tk.END)
                    self.entry_range.insert(0, dim)
                    logger.info(f"CSV dimensions detected: {dim}")
                self.lbl_status.configure(text="✅ Ready")
            except Exception as e:
                logger.error(f"CSV Scan Error: {e}")
            return

        # --- Excel Auto-detect ---
        if not self.workbook: return
        try:
            ws = self.workbook[self.combo_sheet.get()]
            try: dim = ws.calculate_dimension()
            except: dim = "A1:A1"
            self.entry_range.delete(0, tk.END)
            self.entry_range.insert(0, dim)
            logger.info(f"Excel dimensions detected: {dim}")
        except: pass

    def check_queue(self):
        try:
            while True:
                msg = self.msg_queue.get_nowait()
                if msg["type"] == "progress": self.progress.set(msg["value"] / msg["total"]); self.lbl_status.configure(text=f"⏳ {msg.get('file', '...')}")
                elif msg["type"] == "progress_raw": self.progress.set(msg["value"]); self.lbl_status.configure(text=f"⏳ {msg.get('file', '...')}")
                elif msg["type"] == "done": self.out_area.configure(state='normal'); self.out_area.insert("end", msg["result_str"]); self.out_area.see("end"); self.out_area.configure(state='disabled'); self.reset_ui_after_run()
                elif msg["type"] == "error": messagebox.showerror("Error", msg["message"]); self.reset_ui_after_run()
        except queue.Empty: pass
        finally: self.after(100, self.check_queue)

    def toggle_batch_ui(self):
        if self.batch_var.get(): self.btn_browse.configure(text="Select Files..."); self.lbl_file.configure(text=f"{len(self.file_list)} files" if self.file_list else "No files")
        else: self.btn_browse.configure(text="Select File..."); self.lbl_file.configure(text=os.path.basename(self.file_path) if self.file_path else "No file")

    def create_widgets(self):
        f_file = ctk.CTkFrame(self.main_frame, fg_color="transparent"); f_file.pack(fill="x", padx=30, pady=(20, 10))
        h_file = ctk.CTkFrame(f_file, fg_color="transparent"); h_file.pack(fill="x", pady=(0, 5))
        ctk.CTkLabel(h_file, text="Files", font=self.heading_font, text_color=self.text_primary).pack(side="left")
        ctk.CTkButton(h_file, text="⚙️ Settings", font=self.ui_font, fg_color="transparent", text_color=self.text_secondary, hover_color=self.log_bg, width=80, height=25, command=self.open_settings).pack(side="right", padx=(10, 0))
        self.batch_var = tk.BooleanVar(value=False); ctk.CTkSwitch(h_file, text="Batch", font=self.ui_font, variable=self.batch_var, progress_color=self.accent_color, command=self.toggle_batch_ui).pack(side="right")
        sel_bar = ctk.CTkFrame(f_file, fg_color="transparent"); sel_bar.pack(fill="x")
        self.btn_browse = ctk.CTkButton(sel_bar, text="Select File...", font=self.ui_font, fg_color=self.btn_bg, border_width=1, border_color=self.border_color, text_color=self.text_primary, hover_color=self.log_bg, corner_radius=self.radius, height=35, command=self.browse_file); self.btn_browse.pack(side="left")
        self.lbl_file = ctk.CTkLabel(sel_bar, text="No file selected", font=self.ui_font, text_color=self.text_secondary); self.lbl_file.pack(side="left", padx=15)
        f_sheet = ctk.CTkFrame(self.main_frame, fg_color="transparent"); f_sheet.pack(fill="x", padx=30, pady=10)
        ctk.CTkLabel(f_sheet, text="Select sheet", font=self.ui_font, text_color=self.text_secondary).pack(anchor="w")
        self.combo_sheet = ctk.CTkOptionMenu(f_sheet, values=["- empty -"], font=self.ui_font, fg_color=self.btn_bg, button_color=self.btn_bg, text_color=self.text_primary, corner_radius=self.radius); self.combo_sheet.pack(fill="x", pady=(5, 0))
        f_scope = ctk.CTkFrame(self.main_frame, fg_color="transparent"); f_scope.pack(fill="x", padx=30, pady=10)
        ctk.CTkLabel(f_scope, text="Scope", font=self.heading_font, text_color=self.text_primary).pack(anchor="w", pady=(0, 5))
        sc_ctrl = ctk.CTkFrame(f_scope, fg_color="transparent"); sc_ctrl.pack(fill="x")
        self.entry_range = ctk.CTkEntry(sc_ctrl, placeholder_text="e.g. A1:Z100", font=self.ui_font, fg_color=self.btn_bg, border_color=self.border_color, text_color=self.text_primary, corner_radius=self.radius, width=150); self.entry_range.pack(side="left", padx=(0, 10))
        ctk.CTkButton(sc_ctrl, text="Auto-detect", font=self.ui_font, fg_color="transparent", border_width=1, border_color=self.border_color, text_color=self.text_primary, hover_color=self.log_bg, corner_radius=self.radius, command=self.scan_range).pack(side="left")
        f_mode = ctk.CTkFrame(self.main_frame, fg_color="transparent"); f_mode.pack(fill="x", padx=30, pady=10)
        self.mode_var = tk.StringVar(value="Word"); ctk.CTkRadioButton(f_mode, text="Words", font=self.ui_font, variable=self.mode_var, value="Word", fg_color=self.accent_color, text_color=self.text_primary).pack(side="left", padx=(0, 15))
        ctk.CTkRadioButton(f_mode, text="Chars", font=self.ui_font, variable=self.mode_var, value="Char", fg_color=self.accent_color, text_color=self.text_primary).pack(side="left", padx=(0, 15)); ctk.CTkRadioButton(f_mode, text="Both", font=self.ui_font, variable=self.mode_var, value="Both", fg_color=self.accent_color, text_color=self.text_primary).pack(side="left")
        act_f = ctk.CTkFrame(self.main_frame, fg_color="transparent"); act_f.pack(fill="x", padx=30, pady=(20, 5))
        self.btn_run = ctk.CTkButton(act_f, text="Start", font=self.heading_font, fg_color=self.accent_color, hover_color=self.btn_hover, text_color=self.btn_text, corner_radius=self.radius, height=45, command=self.run_count); self.btn_run.pack(side="left", fill="x", expand=True, padx=(0, 10))
        self.btn_report = ctk.CTkButton(act_f, text="Full Report", font=self.heading_font, fg_color="transparent", border_width=1, border_color=self.accent_color, text_color=self.text_primary, hover_color=self.log_bg, corner_radius=self.radius, height=45, command=self.run_report); self.btn_report.pack(side="left", fill="x", expand=True, padx=(0, 10))
        self.btn_cancel = ctk.CTkButton(act_f, text="Cancel", font=self.heading_font, fg_color="#FF4B4B", hover_color="#D32F2F", text_color="#FFFFFF", corner_radius=self.radius, height=45, command=self.cancel_count, state="disabled"); self.btn_cancel.pack(side="left")
        self.progress = ctk.CTkProgressBar(self.main_frame, mode="determinate", progress_color=self.accent_color, fg_color=self.border_color, height=6); self.progress.pack(fill="x", padx=35, pady=5); self.progress.set(0)
        self.lbl_status = ctk.CTkLabel(self.main_frame, text="✅ Ready", font=self.ui_font, text_color=self.text_secondary); self.lbl_status.pack(pady=(0, 10))
        f_log = ctk.CTkFrame(self.main_frame, fg_color="transparent"); f_log.pack(fill="both", expand=True, padx=30, pady=(0, 20))
        log_h = ctk.CTkFrame(f_log, fg_color="transparent"); log_h.pack(fill="x", pady=(0, 5))
        ctk.CTkLabel(log_h, text="Activity Log", font=self.ui_font, text_color=self.text_secondary).pack(side="left")
        ctk.CTkButton(log_h, text="Clear", font=self.ui_font, fg_color="transparent", text_color=self.text_secondary, hover_color=self.log_bg, width=40, height=20, command=self.clear_log).pack(side="right")
        ctk.CTkButton(log_h, text="Export", font=self.ui_font, fg_color="transparent", text_color=self.text_secondary, hover_color=self.log_bg, width=40, height=20, command=self.export_log).pack(side="right", padx=(0, 10))
        self.out_area = ctk.CTkTextbox(f_log, fg_color=self.log_bg, font=self.ui_font, text_color=self.text_primary, corner_radius=self.radius, border_width=0); self.out_area.pack(fill="both", expand=True); self.out_area.configure(state='disabled')

if __name__ == "__main__":
    multiprocessing.freeze_support()
    app = ExcelCounterApp()
    app.mainloop()
